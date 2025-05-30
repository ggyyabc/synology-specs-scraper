import os
from datetime import datetime
from openpyxl import load_workbook

import shutil

def backup_file(filepath):
    """创建备份文件（复制而非重命名）"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(filepath)
    backup_path = f"{base}_backup_{timestamp}{ext}"
    
    try:
        # 复制文件保留元数据
        shutil.copy2(filepath, backup_path)
        print(f"成功创建备份: {backup_path}")
        return backup_path
    except Exception as e:
        print(f"备份失败: {str(e)}")
        raise

def is_row_empty(row):
    """检查行是否完全空白"""
    return all(cell.value is None for cell in row)

def clean_empty_rows(filepath):
    """清理Excel文件中的空白行并调整行高"""
    try:
        # 检查文件是否存在
        if not os.path.exists(filepath):
            print(f"错误：文件 {filepath} 不存在")
            return False

        print(f"正在处理文件: {filepath}")

        # 创建备份
        backup_path = backup_file(filepath)
        print(f"已创建备份文件: {backup_path}")

        # 加载工作簿
        wb = load_workbook(filename=filepath)
        
        # 遍历所有工作表
        for sheet in wb.worksheets:
            print(f"处理工作表: {sheet.title}")
            rows_deleted = 0
            
            # 从最后一行开始检查，避免索引问题
            for row_idx in range(sheet.max_row, 0, -1):
                if is_row_empty(sheet[row_idx]):
                    sheet.delete_rows(row_idx)
                    rows_deleted += 1
            
            print(f"  删除空白行: {rows_deleted} 行")
            
            # 调整行高（跳过前两行）
            print("  正在调整行高...")
            for row in sheet.iter_rows(min_row=3):  # 从第三行开始
                max_lines = 1
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # 计算单元格内容行数（每30个字符算一行）
                        lines = len(cell.value) // 30 + 1
                        max_lines = max(max_lines, lines)
                # 设置行高（每行文字15磅，最小行高15磅）
                sheet.row_dimensions[row[0].row].height = max(15, max_lines * 15)

        # 保存修改
        wb.save(filepath)
        print("处理完成，文件已保存")
        return True

    except Exception as e:
        print(f"处理过程中出错: {str(e)}")
        return False

if __name__ == "__main__":
    import sys
    from tkinter import Tk, filedialog
    
    def select_excel_file():
        """弹出文件选择对话框"""
        root = Tk()
        root.withdraw()  # 隐藏主窗口
        filepath = filedialog.askopenfilename(
            title="选择要处理的Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialdir=os.path.dirname(os.path.abspath(__file__))  # 设置为脚本所在目录
        )
        return filepath
    
    # 两种使用方式：命令行参数或文件选择对话框
    if len(sys.argv) >= 2:
        # 命令行参数方式
        excel_file = ' '.join(sys.argv[1:])
        excel_file = excel_file.strip('"\'')
    else:
        # 文件选择对话框方式
        print("请在弹出的窗口中选择Excel文件...")
        excel_file = select_excel_file()
        if not excel_file:
            print("未选择文件，程序退出")
            sys.exit(0)
    
    excel_file = os.path.abspath(excel_file)
    if not os.path.exists(excel_file):
        print(f"错误：文件不存在 - {excel_file}")
        print("请检查：")
        print("1. 文件路径是否正确")
        print("2. 文件名是否完整（包括.xlsx扩展名）")
        print("3. 文件是否在当前目录下")
        sys.exit(1)
    
    try:
        clean_empty_rows(excel_file)
    except Exception as e:
        print(f"处理文件时发生错误: {str(e)}")
        sys.exit(1)
