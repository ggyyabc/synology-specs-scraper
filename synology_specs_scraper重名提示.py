import requests
from bs4 import BeautifulSoup
import pandas as pd
import tkinter as tk
from tkinter import messagebox
import os
import re
from openpyxl.styles import Font
from openpyxl import load_workbook

EXCEL_FILE = "群晖产品资料汇总.xlsx"

def validate_model_number(model):
    """验证产品型号格式
    典型的群晖产品型号格式如：DS3622xs+, RS4021xs+, DS220+, FS6400
    """
    # 基本格式检查 - 放宽格式限制
    pattern = r'^(DS|RS|FS|SA|HD|DVA)\d{3,4}(\+|xs\+|xs|j|play|II)*$'
    if not re.match(pattern, model):
        return False, "产品型号格式不正确。正确格式示例：DS3622xs+, RS4021xs+, DS220+, FS6400"
    return True, ""

def get_product_specs(model):
    # 首先验证产品型号格式
    is_valid, error_message = validate_model_number(model)
    if not is_valid:
        return False, error_message

    # 构建URL - 直接使用原始型号
    base_url = "https://www.synology.cn/zh-cn/products/"
    url = base_url + model + "#specs"
    
    try:
        # 发送请求
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        # 解析HTML
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 提取规格信息
        specs_data = []
        hardware_section_found = False
        last_spec_item = None  # 用于记录上一个规格项
        section_title = ""  # 用于记录当前部分的标题
        
        # 查找所有表格
        tables = soup.find_all('table')
        
        # 遍历所有表格
        for table in tables:
            # 查找表格的前一个标题
            prev_elem = table.find_previous(['h2', 'h3', 'h4', 'h5', 'div'])
            if prev_elem:
                title_text = prev_elem.get_text(strip=True)
                # 只处理硬件相关的表格
                if any(keyword in title_text.lower() for keyword in ['硬件', 'hardware']):
                    hardware_section_found = True
                    section_title = title_text  # 保存标题，但不添加到数据中
                    last_spec_item = None  # 重置上一个规格项
                elif hardware_section_found:
                    # 如果已经处理完硬件部分，就退出循环
                    break
                else:
                    continue
            
            # 如果不在硬件部分，跳过此表格
            if not hardware_section_found:
                continue
            
            # 处理表格内容
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all(['th', 'td'])
                if cells:
                    # 获取规格项（第一列）
                    spec_name = cells[0].get_text(strip=True)
                    
                    # 获取规格值（第二列，如果存在）
                    spec_value = ""
                    if len(cells) > 1:
                        # 检查是否有特殊标记（如勾号）
                        check_mark = cells[1].find('img', alt='✓')
                        if check_mark:
                            spec_value = "✓"
                        else:
                            spec_value = cells[1].get_text(strip=True)
                    
                    # 获取备注（第三列，如果存在）
                    spec_note = ""
                    if len(cells) > 2:
                        spec_note = cells[2].get_text(strip=True)
                    
                    # 只添加非空的规格项
                    if spec_name or spec_value or spec_note:
                        # 如果规格项与上一个相同，则设为空字符串
                        if spec_name == last_spec_item:
                            spec_name = ""
                        elif spec_name:  # 如果是新的非空规格项
                            last_spec_item = spec_name
                        
                        specs_data.append([spec_name, spec_value, spec_note])
        
        if not specs_data:
            return False, f"未找到产品 {model} 的硬件规格信息。URL: {url}"
            
        # 将数据转换为DataFrame
        df = pd.DataFrame(specs_data, columns=['规格项', '规格值', '备注'])
        
        # 保存到Excel，设置列宽自适应和格式
        try:
            if os.path.exists(EXCEL_FILE):
                with pd.ExcelWriter(EXCEL_FILE, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                    # 添加一个空行作为第一行
                    df.to_excel(writer, sheet_name=model, index=False, startrow=1)
                    # 获取当前工作表
                    worksheet = writer.sheets[model]
                    
                    # 创建粗体字体样式
                    bold_font = Font(bold=True)
                    
                    # 设置列标题行格式（第2行）
                    for cell in worksheet[2]:
                        cell.font = bold_font
            else:
                with pd.ExcelWriter(EXCEL_FILE) as writer:
                    # 添加一个空行作为第一行
                    df.to_excel(writer, sheet_name=model, index=False, startrow=1)
                    # 获取当前工作表
                    worksheet = writer.sheets[model]
                    
                    # 创建粗体字体样式
                    bold_font = Font(bold=True)
                    
                    # 设置列标题行格式（第2行）
                    for cell in worksheet[2]:
                        cell.font = bold_font
                    
        except Exception as e:
            return False, f"保存Excel文件时出错: {str(e)}"
        
        return True, f"硬件规格信息已保存到 {EXCEL_FILE} 的 {model} 工作表中"
        
    except requests.exceptions.RequestException as e:
        return False, f"网络请求错误: {str(e)}\nURL: {url}"
    except Exception as e:
        return False, f"发生错误: {str(e)}\nURL: {url}"

def check_model_exists(model):
    """检查产品型号是否已存在于Excel文件中"""
    if not os.path.exists(EXCEL_FILE):
        return False
    try:
        wb = load_workbook(EXCEL_FILE)
        return model in wb.sheetnames
    except Exception:
        return False

class ProductSpecsApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("群晖产品规格查询")
        self.root.geometry("400x250")
        self.setup_ui()
        self.center_window()
        
    def focus_window(self):
        """激活窗口并设置输入框焦点"""
        self.root.lift()  # 将窗口提升到最前
        self.root.attributes('-topmost', True)  # 设置为顶层窗口
        self.root.attributes('-topmost', False)  # 取消顶层窗口
        self.root.focus_force()  # 强制窗口获得焦点
        self.entry.focus_set()  # 设置输入框焦点

    def setup_ui(self):
        # 添加说明标签
        label = tk.Label(self.root, text="请输入群晖产品型号\n注意：型号需严格区分大小写\n示例: DS3622xs+, RS4021xs+, DS220+", pady=10)
        label.pack()
        
        # 创建输入框框架
        input_frame = tk.Frame(self.root)
        input_frame.pack(pady=5)
        
        # 添加输入框标签
        input_label = tk.Label(input_frame, text="产品型号:")
        input_label.pack(side=tk.LEFT, padx=5)
        
        # 添加输入框
        self.entry = tk.Entry(input_frame, width=25)
        self.entry.pack(side=tk.LEFT, padx=5)
        
        # 绑定回车键
        self.entry.bind('<Return>', lambda event: self.on_submit())
        
        # 添加提交按钮
        submit_btn = tk.Button(self.root, text="获取规格", command=self.on_submit)
        submit_btn.pack(pady=10)
        
        # 添加继续查询复选框
        self.continue_var = tk.BooleanVar(value=True)
        continue_cb = tk.Checkbutton(self.root, text="继续查询下一个产品", variable=self.continue_var)
        continue_cb.pack(pady=5)
        
        # 添加已查询产品数量标签
        self.count_label = tk.Label(self.root, text="已查询产品数量: 0")
        self.count_label.pack(pady=5)
        
        # 添加状态标签
        self.status_label = tk.Label(self.root, text="", fg="gray")
        self.status_label.pack(pady=5)
        
        # 初始化计数器
        self.query_count = 0
        
        # 初始化时激活窗口
        self.root.after(100, self.focus_window)
        
    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
    def on_submit(self):
        model = self.entry.get().strip()
        if not model:
            messagebox.showerror("错误", "请输入产品型号")
            self.root.after(100, self.focus_window)  # 确保窗口激活
            return
            
        # 检查产品型号是否已存在
        if check_model_exists(model):
            response = messagebox.askyesno("提示", 
                f"产品型号 {model} 已存在于Excel文件中。\n是否要重新获取并覆盖现有数据？")
            if not response:
                # 如果用户选择不覆盖，则清空输入框准备下一次输入
                self.entry.delete(0, tk.END)
                self.status_label.config(text="操作已取消", fg="blue")
                self.root.after(100, self.focus_window)
                return
            
        # 更新状态标签
        self.status_label.config(text="正在查询...", fg="blue")
        self.root.update()
        
        success, message = get_product_specs(model)
        if success:
            self.query_count += 1
            self.count_label.config(text=f"已查询产品数量: {self.query_count}")
            self.status_label.config(text="查询成功", fg="green")
            messagebox.showinfo("结果", message)
            
            if self.continue_var.get():
                # 清空输入框，准备下一次输入
                self.entry.delete(0, tk.END)
                self.status_label.config(text="请输入下一个产品型号", fg="gray")
                self.root.after(100, self.focus_window)
            else:
                self.root.destroy()
        else:
            self.status_label.config(text="查询失败", fg="red")
            messagebox.showerror("错误", message)
            # 选中输入框中的文本，方便用户直接修改
            self.entry.select_range(0, tk.END)
            self.root.after(100, self.focus_window)
    
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = ProductSpecsApp()
    app.run() 