import requests
from bs4 import BeautifulSoup
import pandas as pd
import tkinter as tk
from tkinter import messagebox
import os
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO
import urllib.parse
import tempfile
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU
from datetime import datetime
from openpyxl.worksheet.hyperlink import Hyperlink

# 版本信息
__version__ = "1.4"
__author__ = "Claude"
__update_notes__ = """
V1.4更新说明：
1. 修复了扩展设备和网卡的规格获取问题
   - 支持RX1225RP、RX1223RP等带RP后缀的型号
   - 支持RXD1219sas等新一代扩展设备
   - 支持E10G22-T1-Mini等迷你网卡
   - 支持M2D20等M.2扩充卡

2. 改进了产品图片获取功能
   - 修复了网卡和PCIe设备的图片下载
   - 针对不同类型产品使用不同的图片参数
   - 自动尝试多个图片URL格式
   - 优化了图片下载失败的重试机制

3. 优化了排序功能
   - 移除了不安全的超链接排序
   - 改用程序界面按钮进行排序
   - 添加了升序/降序切换按钮
   - 优化了排序状态显示
"""

EXCEL_FILE = "群晖产品资料汇总.xlsx"
IMAGES_DIR = "产品图片"  # 图片保存目录
SUMMARY_SHEET = "产品汇总表"  # 汇总表名称

# 定义样式常量
HEADER_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BORDER_STYLE = Side(style='thin', color="000000")
NORMAL_BORDER = Border(left=BORDER_STYLE, right=BORDER_STYLE, top=BORDER_STYLE, bottom=BORDER_STYLE)

def validate_model_number(model):
    """验证产品型号格式
    支持的格式示例：
    存储扩充设备：
    - RX1217sas, RX1217, RX418 (机架式扩展设备)
    - DX1215, DX517 (桌面式扩展设备)
    - FX2421, FX2421rp (全闪存扩展设备)
    - RXD1219sas (新一代扩展设备)
    - RX1223RP, RX1225RP (带冗余电源的扩展设备)
    
    PCIe 扩充卡：
    - E10G18-T2, E10G18-T1
    - E25G21-F2
    - M2D20, M2D18
    - FXC17, FXC18
    - E10G22-T1-Mini (迷你网卡)
    """
    # 基本格式检查 - 支持配件产品线
    patterns = [
        # 存储扩充设备
        r'^RX\d{3,4}(sas|RP)?$',  # RX系列基本型号
        r'^RXD\d{4}sas$',         # RXD系列
        r'^DX\d{3,4}$',           # DX系列
        r'^FX\d{4}(rp)?$',        # FX系列
        
        # PCIe 扩充卡
        r'^[A-Z]\d{2}[A-Z]\d{2}-[A-Z]\d{1,2}(-Mini)?$',  # 新格式：E10G22-T1-Mini
        r'^[A-Z]\d{2}[A-Z]\d{2}-[A-Z]\d{1,2}$',          # 新格式：E25G21-F2
        r'^[A-Z]\d{1,2}[A-Z]\d{2}(-T\d)?$',              # 旧格式：E10G18-T2
        r'^[A-Z]\d{2}[A-Z]\d{2}$',                        # M2D20, FXC18等
        
        # NAS/SAN系列保持不变
        r'^(DS|RS|FS|SA|HD|DVA|UC)\d{3,4}(RP)?(xs\+|xs|\+|slim|play|j|II|D)?$'
    ]
    
    for pattern in patterns:
        if re.match(pattern, model):
            return True, ""
    
    return False, """产品型号格式不正确。正确格式示例：
存储扩充设备：
- RX1217sas, RX1223RP, RX1225RP
- RXD1219sas
- DX1215, DX517
- FX2421, FX2421rp

PCIe 扩充卡：
- E10G22-T1-Mini
- E25G21-F2
- E10G18-T2
- M2D20, FXC18"""

def calculate_row_height(row):
    """计算行高
    根据单元格内容和换行数量计算合适的行高
    标准行高为15，每多一行增加15
    """
    max_lines = 1
    for cell in row:
        if cell.value:
            # 计算文本换行后的行数
            text = str(cell.value)
            # 获取单元格宽度（以字符为单位）
            col_width = cell.parent.column_dimensions[get_column_letter(cell.column)].width
            
            # 如果启用了自动换行，计算实际行数
            if cell.alignment and cell.alignment.wrap_text:
                # 预估每行能容纳的字符数（考虑中文字符）
                chars_per_line = int(col_width / 1.5)  # 假设每个中文字符宽度为1.5
                if chars_per_line > 0:
                    # 计算需要的行数
                    lines = len(text) / chars_per_line
                    max_lines = max(max_lines, int(lines) + 1)
            else:
                # 未启用自动换行时，只计算手动换行符
                lines = text.count('\n') + 1
                max_lines = max(max_lines, lines)
    
    # 基础行高15，每行增加15
    return max(20, 15 * max_lines)

def ensure_dir(directory):
    """确保目录存在，如果不存在则创建"""
    if not os.path.exists(directory):
        os.makedirs(directory)

def make_background_transparent(img):
    """将图片的白色背景转换为透明"""
    # 转换图片为RGBA模式（支持透明通道）
    img = img.convert("RGBA")
    data = img.getdata()
    
    # 创建新的像素数据，将接近白色的像素转换为透明
    new_data = []
    for item in data:
        # 检查像素是否接近白色（RGB值都大于240）
        if item[0] > 240 and item[1] > 240 and item[2] > 240:
            # 将白色像素转换为完全透明
            new_data.append((255, 255, 255, 0))
        else:
            new_data.append(item)
    
    # 更新图片数据
    img.putdata(new_data)
    return img

def download_and_resize_image(model):
    """下载并调整产品图片大小，同时保存到本地"""
    # 确保图片目录存在
    ensure_dir(IMAGES_DIR)
    
    # 构建图片URL
    encoded_model = urllib.parse.quote(model)
    
    # 根据产品类型选择不同的图片URL
    if re.match(r'^M2D\d{2}', model):  # M2D系列
        image_url = f"https://www.synology.cn/api/products/getPhoto?product={encoded_model}&type=img&sort=1"
    elif re.match(r'^[A-Z]\d{2}[A-Z]', model):  # 网卡和其他PCIe设备
        image_url = f"https://www.synology.cn/api/products/getPhoto?product={encoded_model}&type=img&sort=0"
    else:  # NAS和扩展设备
        image_url = f"https://www.synology.cn/api/products/getPhoto?product={encoded_model}&type=img&sort=2"
    
    try:
        # 下载图片
        response = requests.get(image_url, timeout=10)
        
        # 检查响应状态码
        if response.status_code != 200:
            print(f"下载图片失败，状态码: {response.status_code}")
            # 如果第一次尝试失败，尝试其他sort参数
            sort_values = ['0', '1', '2']
            current_sort = image_url.split('sort=')[1]
            for sort_value in sort_values:
                if sort_value != current_sort:
                    new_url = image_url.replace(f'sort={current_sort}', f'sort={sort_value}')
                    try:
                        response = requests.get(new_url, timeout=10)
                        if response.status_code == 200:
                            image_url = new_url
                            break
                    except:
                        continue
            
            if response.status_code != 200:
                return None
            
        # 检查内容类型
        content_type = response.headers.get('content-type', '')
        if not content_type.startswith('image/'):
            print(f"返回的内容不是图片: {content_type}")
            return None
            
        # 检查内容长度
        if len(response.content) < 100:
            print("返回的图片数据异常")
            return None
        
        try:
            # 使用PIL打开图片并验证
            img = PILImage.open(BytesIO(response.content))
            img.verify()
            
            # 重新打开图片（verify后需要重新打开）
            img = PILImage.open(BytesIO(response.content))
            
            # 检查图片尺寸
            if img.size[0] < 10 or img.size[1] < 10:
                print("图片尺寸异常")
                return None
            
            # 保存原始图片到本地
            original_path = os.path.join(IMAGES_DIR, f"{model}.png")
            img.save(original_path, format='PNG')
            print(f"原始图片已保存到: {original_path}")
            
            # 计算等比例缩放后的高度（Excel中显示用）
            width = 140  # 调整Excel中显示的图片宽度
            ratio = width / float(img.size[0])
            height = int(float(img.size[1]) * ratio)
            
            # 调整图片大小
            img_resized = img.resize((width, height), PILImage.Resampling.LANCZOS)
            
            # 将调整后的图片保存到内存中
            img_byte_arr = BytesIO()
            img_resized.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            
            return img_byte_arr, height  # 返回图片数据和高度
            
        except (IOError, OSError) as e:
            print(f"处理图片时出错: {str(e)}")
            return None
            
    except requests.exceptions.RequestException as e:
        print(f"下载图片时出错: {str(e)}")
        return None
    except Exception as e:
        print(f"发生未知错误: {str(e)}")
        return None

def format_worksheet(worksheet, df, model):
    """设置工作表格式"""
    # 设置第一行高度为固定值
    worksheet.row_dimensions[1].height = 120
    
    # 下载并插入产品图片
    img_result = download_and_resize_image(model)
    if img_result:
        img_data, img_height = img_result
        try:
            # 在A1单元格插入图片
            img = Image(img_data)
            # 设置图片位置（A1单元格内）
            img.anchor = 'A1'
            worksheet.add_image(img)
        except Exception as e:
            print(f"插入图片时出错: {str(e)}")
    
    # 先取消所有合并的单元格
    # 创建合并范围的列表副本
    merged_ranges = list(worksheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        worksheet.unmerge_cells(str(merged_range))
    
    # 设置标题行格式（第1行）
    # 先设置A1单元格的值和格式
    title_cell = worksheet['A1']
    title_cell.value = f'群晖{model} 硬件规格'
    title_cell.font = Font(
        name='微软雅黑',  # 设置字体为微软雅黑
        size=24,         # 设置字号为24
        bold=True,       # 加粗
        italic=True      # 斜体
    )
    title_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    
    # 合并A1:C1单元格
    worksheet.merge_cells('A1:C1')
    
    # 在E1单元格添加返回按钮
    back_cell = worksheet['E1']
    back_cell.value = "返回产品汇总表"
    
    # 创建相对引用的超链接
    back_cell.hyperlink = Hyperlink(
        display="返回产品汇总表",
        ref="E1",  # 当前单元格的引用
        location=f"'{SUMMARY_SHEET}'!A1",  # 目标位置
        target="#'产品汇总表'!A1"  # 使用#前缀的相对路径
    )
    
    back_cell.font = Font(
        name='微软雅黑',
        size=18,
        color="0563C1",
        underline="single",
        bold=True
    )
    back_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列标题行格式（第2行）
    header_row = worksheet[2]
    for cell in header_row[:3]:  # 只处理前三列
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = NORMAL_BORDER
    
    # 设置固定列宽
    worksheet.column_dimensions['A'].width = 20  # 调整A列宽度
    worksheet.column_dimensions['B'].width = 30  # 规格项列
    worksheet.column_dimensions['C'].width = 73  # 规格值列
    worksheet.column_dimensions['E'].width = 25  # 返回按钮列
    
    # 设置表头行高
    worksheet.row_dimensions[2].height = 20  # 表头行高固定
    
    # 获取所有大类（第一列非空值）
    categories = []
    last_category = None
    category_rows = []  # 存储每个大类的起始行号
    
    # 处理数据行
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, max_row=worksheet.max_row), start=3):
        cell_value = row[0].value
        if cell_value:  # 如果第一列有值，说明是新的大类
            categories.append(cell_value)
            category_rows.append(row_idx)
            last_category = cell_value
            # 设置大类单元格格式
            row[0].font = Font(bold=True)
            row[0].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        else:
            # 对于大类下的子项，缩进第二列
            if row[1].value:
                row[1].alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
        
        # 设置规格值列的对齐方式和自动换行
        if row[2].value:
            row[2].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 只为前三列添加边框
        for cell in row[:3]:
            cell.border = NORMAL_BORDER
            # 确保所有单元格都启用自动换行
            if cell.alignment:
                new_alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=True,
                    indent=cell.alignment.indent
                )
                cell.alignment = new_alignment
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 设置行高
        row_height = calculate_row_height(row)
        worksheet.row_dimensions[row_idx].height = row_height
    
    # 设置打印相关属性
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.page_setup.fitToPage = True
    worksheet.page_setup.fitToHeight = False
    worksheet.page_setup.fitToWidth = 1

def create_or_update_summary_sheet(workbook, model=None, sort_by=None, sort_ascending=None):
    """创建或更新产品汇总表"""
    # 如果汇总表不存在，创建它
    if SUMMARY_SHEET not in workbook.sheetnames:
        summary_sheet = workbook.create_sheet(SUMMARY_SHEET, 0)  # 在最前面创建
        
        # 添加排序按钮
        summary_sheet.cell(row=1, column=6, value="▼").font = Font(bold=True)  # F1: 降序按钮
        summary_sheet.cell(row=1, column=7, value="▲").font = Font(bold=True)  # G1: 升序按钮
        summary_sheet.column_dimensions['F'].width = 5  # 设置按钮列宽
        summary_sheet.column_dimensions['G'].width = 5
        
        # 设置表头
        headers = ['序号', '产品型号', '添加时间', '备注']
        for col, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = NORMAL_BORDER
        
        # 设置列宽
        summary_sheet.column_dimensions['A'].width = 8   # 序号
        summary_sheet.column_dimensions['B'].width = 20  # 产品型号
        summary_sheet.column_dimensions['C'].width = 20  # 添加时间
        summary_sheet.column_dimensions['D'].width = 30  # 备注
        
        # 冻结首行
        summary_sheet.freeze_panes = 'A2'
        
        # 添加排序说明
        sort_note = summary_sheet.cell(row=1, column=5)  # E1单元格
        sort_note.value = '使用程序界面的排序按钮进行排序'
        sort_note.font = Font(color="808080", italic=True)  # 灰色斜体
        summary_sheet.column_dimensions['E'].width = 35  # 设置说明列宽
        
    else:
        summary_sheet = workbook[SUMMARY_SHEET]
    
    # 获取所有产品工作表（排除汇总表）
    product_sheets = [sheet for sheet in workbook.sheetnames if sheet != SUMMARY_SHEET]
    
    # 收集现有数据（包括时间和备注）
    existing_data = {}
    for row in range(2, summary_sheet.max_row + 1):
        product_name = summary_sheet.cell(row=row, column=2).value
        if product_name:
            existing_data[product_name] = {
                'time': summary_sheet.cell(row=row, column=3).value,
                'note': summary_sheet.cell(row=row, column=4).value
            }
    
    # 清空现有数据（保留表头）
    for row in range(2, summary_sheet.max_row + 1):
        for col in range(1, 5):
            cell = summary_sheet.cell(row=row, column=col)
            cell.value = None
            cell.hyperlink = None  # 清除超链接
    
    # 收集所有产品数据
    products_data = []
    for sheet_name in product_sheets:
        # 如果是当前添加的产品，使用当前时间
        is_current_product = (sheet_name == model)
        
        # 获取或设置时间
        if is_current_product:
            time_value = datetime.now().strftime('%Y-%m-%d %H:%M')
        elif sheet_name in existing_data:
            time_value = existing_data[sheet_name]['time']
        else:
            time_value = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        # 获取备注
        note = existing_data.get(sheet_name, {}).get('note', '')
        
        products_data.append({
            'name': sheet_name,
            'time': time_value,
            'note': note
        })
    
    # 排序数据
    if sort_by == 'name':
        products_data.sort(key=lambda x: x['name'], reverse=not sort_ascending)
    elif sort_by == 'time':
        products_data.sort(key=lambda x: x['time'], reverse=not sort_ascending)
    else:
        # 默认按产品型号排序
        products_data.sort(key=lambda x: x['name'])
    
    # 更新表头文本
    model_header = summary_sheet['B1']
    time_header = summary_sheet['C1']
    
    if sort_by == 'name':
        model_header.value = f"产品型号 {'▲' if sort_ascending else '▼'}"
        time_header.value = "添加时间"
    elif sort_by == 'time':
        model_header.value = "产品型号"
        time_header.value = f"添加时间 {'▲' if sort_ascending else '▼'}"
    else:
        model_header.value = "产品型号"
        time_header.value = "添加时间"
    
    # 设置字体样式
    model_header.font = Font(bold=True)
    time_header.font = Font(bold=True)
    
    # 重新添加所有产品
    for idx, product in enumerate(products_data, 1):
        # 序号
        summary_sheet.cell(row=idx+1, column=1, value=idx)
        
        # 产品型号（添加超链接）
        cell = summary_sheet.cell(row=idx+1, column=2)
        cell.value = product['name']
        cell.hyperlink = Hyperlink(
            display=product['name'],
            ref=f"A{idx+1}",
            location=f"'{product['name']}'!A1",
            target=f"#{product['name']}!A1"
        )
        cell.font = Font(color="0563C1", underline="single")
        
        # 添加时间
        summary_sheet.cell(row=idx+1, column=3, value=product['time'])
        
        # 添加备注
        summary_sheet.cell(row=idx+1, column=4, value=product['note'])
        
        # 设置单元格边框和对齐方式
        for col in range(1, 5):
            cell = summary_sheet.cell(row=idx+1, column=col)
            cell.border = NORMAL_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 保护工作表，只允许点击超链接和编辑备注
    summary_sheet.protection.sheet = True
    summary_sheet.protection.enable()

def update_all_summary():
    """更新所有已有产品的汇总表"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return False, "Excel文件不存在"
        
        try:
            workbook = load_workbook(EXCEL_FILE)
        except Exception as e:
            return False, f"无法打开Excel文件: {str(e)}"
        
        create_or_update_summary_sheet(workbook)
        
        try:
            workbook.save(EXCEL_FILE)
        except Exception as e:
            return False, f"保存Excel文件时出错: {str(e)}"
        
        return True, "汇总表更新成功"
        
    except Exception as e:
        return False, f"更新汇总表时出错: {str(e)}"

def get_product_specs(model):
    # 首先验证产品型号格式
    is_valid, error_message = validate_model_number(model)
    if not is_valid:
        return False, error_message

    # 构建URL - 直接使用完整路径
    base_url = "https://www.synology.cn/zh-cn/products/"
    
    # 对于M2D系列和网卡，使用特殊的URL路径
    if re.match(r'^M2D\d{2}', model) or re.match(r'^[A-Z]\d{2}[A-Z]', model):
        base_url = "https://www.synology.cn/zh-cn/products/M2_PCIe_Card/"
    
    url = base_url + model + "#specs"
    
    try:
        # 发送请求
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers)
        
        # 如果主URL返回404，尝试其他可能的URL
        if response.status_code == 404:
            # 尝试不同的URL路径
            alternate_urls = [
                f"https://www.synology.cn/zh-cn/products/{model}",  # 无#specs后缀
                f"https://www.synology.cn/zh-cn/products/M2_PCIe_Card/{model}",  # M2D系列路径
                f"https://www.synology.cn/zh-cn/products/network/{model}",  # 网卡路径
                f"https://www.synology.cn/zh-cn/products/PCIe_Card/{model}"  # PCIe卡通用路径
            ]
            
            for alt_url in alternate_urls:
                try:
                    response = requests.get(alt_url, headers=headers)
                    if response.status_code == 200:
                        url = alt_url  # 更新为成功的URL
                        break
                except:
                    continue
        
        response.raise_for_status()
        
        # 解析HTML
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 提取规格信息
        specs_data = []
        hardware_section_found = False
        last_spec_item = None  # 用于记录上一个规格项
        
        # 查找所有表格
        tables = soup.find_all('table')
        
        # 遍历所有表格
        for table in tables:
            # 查找表格的前一个标题
            prev_elem = table.find_previous(['h2', 'h3', 'h4', 'h5', 'div'])
            if prev_elem:
                title_text = prev_elem.get_text(strip=True)
                # 处理规格相关的表格
                if any(keyword in title_text.lower() for keyword in ['硬件', 'hardware', '规格', 'specifications']):
                    hardware_section_found = True
                    section_title = title_text
                    last_spec_item = None  # 重置上一个规格项
                elif hardware_section_found:
                    # 如果已经处理完规格部分，就退出循环
                    break
                else:
                    continue
            
            # 如果不在规格部分，跳过此表格
            if not hardware_section_found:
                continue
            
            # 处理表格内容
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all(['th', 'td'])
                if cells:
                    # 获取规格项（第一列）
                    spec_name = cells[0].get_text(strip=True)
                    
                    # 获取规格值（第二列，如果存在）
                    spec_value = ""
                    if len(cells) > 1:
                        # 检查是否有特殊标记（如勾号）
                        check_mark = cells[1].find('img', alt='✓')
                        if check_mark:
                            spec_value = "✓"
                        else:
                            spec_value = cells[1].get_text(strip=True)
                    
                    # 获取备注（第三列，如果存在）
                    spec_note = ""
                    if len(cells) > 2:
                        spec_note = cells[2].get_text(strip=True)
                    
                    # 只添加非空的规格项
                    if spec_name or spec_value or spec_note:
                        # 如果规格项与上一个相同，则设为空字符串
                        if spec_name == last_spec_item:
                            spec_name = ""
                        elif spec_name:  # 如果是新的非空规格项
                            last_spec_item = spec_name
                        
                        specs_data.append([spec_name, spec_value, spec_note])
        
        if not specs_data:
            return False, f"未找到产品 {model} 的规格信息。URL: {url}"
            
        # 将数据转换为DataFrame
        df = pd.DataFrame(specs_data, columns=['规格项', '规格值', '技术指标'])
        
        # 保存到Excel，设置格式
        try:
            # 如果文件存在且可能损坏，先尝试创建备份
            if os.path.exists(EXCEL_FILE):
                try:
                    # 尝试打开现有文件以验证其完整性
                    wb = load_workbook(EXCEL_FILE)
                    wb.close()
                except Exception as e:
                    # 如果文件损坏，创建备份并创建新文件
                    backup_file = f"{EXCEL_FILE}.bak"
                    if os.path.exists(backup_file):
                        os.remove(backup_file)
                    os.rename(EXCEL_FILE, backup_file)
                    print(f"原文件已损坏，已创建备份：{backup_file}")
            
            # 创建新的Excel文件或追加到现有文件
            if os.path.exists(EXCEL_FILE):
                # 使用with语句确保文件正确关闭
                with pd.ExcelWriter(EXCEL_FILE, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                    # 添加一个空行作为第一行，从第二行开始写入数据
                    df.to_excel(writer, sheet_name=model, index=False, startrow=1)
                    # 获取当前工作表
                    workbook = writer.book
                    worksheet = writer.sheets[model]
                    
                    # 合并第一行单元格并添加标题
                    worksheet.merge_cells('A1:C1')
                    worksheet['A1'] = f'群晖{model} 硬件规格'
                    
                    # 应用格式化
                    format_worksheet(worksheet, df, model)
                    
                    # 更新汇总表
                    create_or_update_summary_sheet(workbook, model)
            else:
                # 创建新文件
                with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
                    # 添加一个空行作为第一行，从第二行开始写入数据
                    df.to_excel(writer, sheet_name=model, index=False, startrow=1)
                    # 获取当前工作表
                    workbook = writer.book
                    worksheet = writer.sheets[model]
                    
                    # 合并第一行单元格并添加标题
                    worksheet.merge_cells('A1:C1')
                    worksheet['A1'] = f'群晖{model} 硬件规格'
                    
                    # 应用格式化
                    format_worksheet(worksheet, df, model)
                    
                    # 创建汇总表
                    create_or_update_summary_sheet(workbook, model)
                    
        except Exception as e:
            error_msg = str(e)
            # 如果是文件被占用的错误，给出更友好的提示
            if "Permission denied" in error_msg or "being used by another process" in error_msg:
                return False, f"无法保存Excel文件，请确保文件未被其他程序打开: {error_msg}"
            return False, f"保存Excel文件时出错: {error_msg}"
        
        return True, f"规格信息已保存到 {EXCEL_FILE} 的 {model} 工作表中"
        
    except requests.exceptions.RequestException as e:
        return False, f"网络请求错误: {str(e)}\nURL: {url}"
    except Exception as e:
        return False, f"发生错误: {str(e)}\nURL: {url}"

def check_model_exists(model):
    """检查产品型号是否已存在于Excel文件中"""
    if not os.path.exists(EXCEL_FILE):
        return False
    try:
        wb = load_workbook(EXCEL_FILE)
        return model in wb.sheetnames
    except Exception:
        return False

class ProductSpecsApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title(f"群晖产品规格查询 V{__version__}")
        self.root.geometry("400x400")  # 增加窗口高度
        self.setup_ui()
        self.center_window()
        
        # 初始化排序状态
        self.sort_state = {'by': None, 'ascending': True}
        
        # 监听Excel文件变化
        self.watch_excel_file()
        
    def watch_excel_file(self):
        """监听Excel文件的变化，检查是否有排序请求"""
        if os.path.exists(EXCEL_FILE):
            try:
                workbook = load_workbook(EXCEL_FILE)
                if SUMMARY_SHEET in workbook.sheetnames:
                    sheet = workbook[SUMMARY_SHEET]
                    
                    # 检查B1和C1单元格的超链接
                    for cell in [sheet['B1'], sheet['C1']]:
                        if cell.hyperlink and cell.hyperlink.target:
                            if cell.hyperlink.target == "sort_name":
                                self.on_sort('name', True)
                                break
                            elif cell.hyperlink.target == "sort_time":
                                self.on_sort('time', True)
                                break
                workbook.close()
            except Exception:
                pass
        
        # 每500毫秒检查一次
        self.root.after(500, self.watch_excel_file)
    
    def focus_window(self):
        """激活窗口并设置输入框焦点"""
        self.root.lift()  # 将窗口提升到最前
        self.root.attributes('-topmost', True)  # 设置为顶层窗口
        self.root.attributes('-topmost', False)  # 取消顶层窗口
        self.root.focus_force()  # 强制窗口获得焦点
        self.entry.focus_set()  # 设置输入框焦点

    def setup_ui(self):
        # 添加说明标签
        label = tk.Label(self.root, text="请输入群晖产品型号\n注意：型号需严格区分大小写\n示例: DS3622xs+, RS4021xs+, DS220+", pady=10)
        label.pack()
        
        # 创建输入框框架
        input_frame = tk.Frame(self.root)
        input_frame.pack(pady=5)
        
        # 添加输入框标签
        input_label = tk.Label(input_frame, text="产品型号:")
        input_label.pack(side=tk.LEFT, padx=5)
        
        # 添加输入框
        self.entry = tk.Entry(input_frame, width=25)
        self.entry.pack(side=tk.LEFT, padx=5)
        
        # 绑定回车键
        self.entry.bind('<Return>', lambda event: self.on_submit())
        
        # 添加提交按钮
        submit_btn = tk.Button(self.root, text="获取规格", command=self.on_submit)
        submit_btn.pack(pady=10)
        
        # 添加更新汇总表按钮
        update_summary_btn = tk.Button(self.root, text="更新汇总表", command=self.on_update_summary)
        update_summary_btn.pack(pady=5)
        
        # 添加继续查询复选框
        self.continue_var = tk.BooleanVar(value=True)
        continue_cb = tk.Checkbutton(self.root, text="继续查询下一个产品", variable=self.continue_var)
        continue_cb.pack(pady=5)
        
        # 添加已查询产品数量标签
        self.count_label = tk.Label(self.root, text="已查询产品数量: 0")
        self.count_label.pack(pady=5)
        
        # 添加状态标签
        self.status_label = tk.Label(self.root, text="", fg="gray")
        self.status_label.pack(pady=5)
        
        # 初始化计数器
        self.query_count = 0
        
        # 添加排序区域标题
        sort_title = tk.Label(self.root, text="排序选项", font=("", 10, "bold"))
        sort_title.pack(pady=(10, 5))
        
        # 添加排序按钮框架
        sort_frame = tk.Frame(self.root)
        sort_frame.pack(pady=5)
        
        # 创建产品型号排序按钮
        model_frame = tk.LabelFrame(sort_frame, text="按产品型号", padx=5, pady=5)
        model_frame.pack(side=tk.LEFT, padx=10)
        
        tk.Button(model_frame, text="▲ 升序", command=lambda: self.on_sort('name', True)).pack(side=tk.LEFT, padx=2)
        tk.Button(model_frame, text="▼ 降序", command=lambda: self.on_sort('name', False)).pack(side=tk.LEFT, padx=2)
        
        # 创建更新时间排序按钮
        time_frame = tk.LabelFrame(sort_frame, text="按更新时间", padx=5, pady=5)
        time_frame.pack(side=tk.LEFT, padx=10)
        
        tk.Button(time_frame, text="▲ 升序", command=lambda: self.on_sort('time', True)).pack(side=tk.LEFT, padx=2)
        tk.Button(time_frame, text="▼ 降序", command=lambda: self.on_sort('time', False)).pack(side=tk.LEFT, padx=2)
        
        # 初始化时激活窗口
        self.root.after(100, self.focus_window)
        
    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
    def on_submit(self):
        model = self.entry.get().strip()
        if not model:
            messagebox.showerror("错误", "请输入产品型号")
            self.root.after(100, self.focus_window)  # 确保窗口激活
            return
            
        # 检查产品型号是否已存在
        if check_model_exists(model):
            response = messagebox.askyesno("提示", 
                f"产品型号 {model} 已存在于Excel文件中。\n是否要重新获取并覆盖现有数据？")
            if not response:
                # 如果用户选择不覆盖，则清空输入框准备下一次输入
                self.entry.delete(0, tk.END)
                self.status_label.config(text="操作已取消", fg="blue")
                self.root.after(100, self.focus_window)
                return
            
        # 更新状态标签
        self.status_label.config(text="正在查询...", fg="blue")
        self.root.update()
        
        success, message = get_product_specs(model)
        if success:
            self.query_count += 1
            self.count_label.config(text=f"已查询产品数量: {self.query_count}")
            self.status_label.config(text="查询成功", fg="green")
            messagebox.showinfo("结果", message)
            
            if self.continue_var.get():
                # 清空输入框，准备下一次输入
                self.entry.delete(0, tk.END)
                self.status_label.config(text="请输入下一个产品型号", fg="gray")
                self.root.after(100, self.focus_window)
            else:
                self.root.destroy()
        else:
            self.status_label.config(text="查询失败", fg="red")
            messagebox.showerror("错误", message)
            # 选中输入框中的文本，方便用户直接修改
            self.entry.select_range(0, tk.END)
            self.root.after(100, self.focus_window)
    
    def on_update_summary(self):
        """更新汇总表按钮的点击事件处理"""
        success, message = update_all_summary()
        if success:
            self.status_label.config(text="汇总表更新成功", fg="green")
            messagebox.showinfo("成功", message)
        else:
            self.status_label.config(text="汇总表更新失败", fg="red")
            messagebox.showerror("错误", message)
        self.root.after(100, self.focus_window)
    
    def on_sort(self, sort_by, ascending):
        """排序按钮点击事件处理"""
        if os.path.exists(EXCEL_FILE):
            try:
                # 应用排序
                workbook = load_workbook(EXCEL_FILE)
                create_or_update_summary_sheet(
                    workbook,
                    sort_by=sort_by,
                    sort_ascending=ascending
                )
                workbook.save(EXCEL_FILE)
                
                self.status_label.config(
                    text=f"已按{'产品型号' if sort_by == 'name' else '添加时间'}{'升序' if ascending else '降序'}排序",
                    fg="green"
                )
            except Exception as e:
                self.status_label.config(text=f"排序失败: {str(e)}", fg="red")
                messagebox.showerror("错误", f"排序失败: {str(e)}")
        else:
            self.status_label.config(text="Excel文件不存在", fg="red")
            messagebox.showerror("错误", "Excel文件不存在")
        
        self.root.after(100, self.focus_window)
    
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = ProductSpecsApp()
    app.run() 