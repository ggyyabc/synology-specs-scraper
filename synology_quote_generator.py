import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os

# 版本信息
__version__ = "1.2"
__author__ = "Claude"

# 常量定义
SPECS_FILE = "群晖产品资料汇总.xlsx"  # 产品规格文件
QUOTE_DIR = "客户报价单文件夹"  # 报价单输出目录
CUSTOMER_INFO_DIR = "客户信息文件夹"  # 客户信息记录目录
CUSTOMER_INFO_FILE = os.path.join(CUSTOMER_INFO_DIR, "客户信息表.xlsx")  # 客户信息记录文件

# 产品类型定义
PRODUCT_CATEGORIES = {
    "NAS设备": {
        "DS系列": ["DS1621+", "DS1821+", "DS2422+", "DS3622xs+"],
        "RS系列": ["RS1221+", "RS3621xs+", "RS4021xs+"],
        "FS系列": ["FS2500", "FS3600", "FS6400"],
        "SA系列": ["SA3200D"],
        "UC系列": ["UC3200", "UC3400"],
    },
    "存储扩充设备": {
        "RX系列": ["RX418", "RX1217", "RX1223RP", "RX1225RP"],
        "DX系列": ["DX517"],
        "FX系列": ["FX2421"],
        "RXD系列": ["RXD1219sas"],
    },
    "PCIe扩充卡": {
        "网卡": ["E10G18-T1", "E10G22-T1-Mini"],
        "M.2转接卡": ["M2D20", "M2D18"],
    },
    "硬盘": {
        "企业级硬盘": [],  # 将从Excel文件中读取
        "数据中心硬盘": [],
        "监控级硬盘": [],
    },
    "内存": {
        "ECC内存": [],  # 将从Excel文件中读取
        "非ECC内存": [],
    }
}

# Excel样式定义
HEADER_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
BORDER_STYLE = Side(style='thin', color="000000")
NORMAL_BORDER = Border(left=BORDER_STYLE, right=BORDER_STYLE, top=BORDER_STYLE, bottom=BORDER_STYLE)

def ensure_directories():
    """确保必要的目录存在"""
    directories = [QUOTE_DIR, CUSTOMER_INFO_DIR]
    for directory in directories:
            os.makedirs(directory, exist_ok=True)
            print(f"确保目录存在：{directory}")

class QuoteGenerator:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title(f"云南金海维-群晖产品报价单生成器 V{__version__}")
        self.root.geometry("1200x700")
        
        # 确保必要的目录存在
        ensure_directories()
        
        # 初始化产品数据
        self.available_products = []
        self.product_categories = {}
        
        # 加载现有客户数据
        self.existing_customers = self.load_existing_customers()
        
        # 加载产品数据
        self.load_product_data()
        
        # 初始化UI
        self.setup_ui()
        self.center_window()
        
        # 报价项目列表
        self.quote_items = []
    
    def load_existing_customers(self):
        """加载现有客户信息"""
        customers = {}
        if os.path.exists(CUSTOMER_INFO_FILE):
            try:
                df = pd.read_excel(CUSTOMER_INFO_FILE)
                for _, row in df.iterrows():
                    customer_name = str(row.get('客户名称', '')).strip()
                    if customer_name:
                        customers[customer_name] = {
                            'contact': str(row.get('联系人', '')).strip(),
                            'phone': str(row.get('联系电话', '')).strip()
                        }
                print(f"已加载 {len(customers)} 个现有客户信息")
            except Exception as e:
                print(f"加载客户信息时出错：{str(e)}")
        return customers
    
    def load_product_data(self):
        """加载产品数据"""
        try:
            if not os.path.exists(SPECS_FILE):
                messagebox.showerror("错误", f"未找到产品规格文件：{SPECS_FILE}")
                return
            
            print("\n开始加载产品数据...")
            
            # 读取Excel文件中的所有工作表名（产品型号）
            wb = load_workbook(SPECS_FILE, read_only=True)
            self.available_products = [sheet for sheet in wb.sheetnames 
                                    if sheet not in ["产品汇总表", "查询表格", "硬盘分类查询表格", "配件分类查询表格"]]
            print(f"找到 {len(self.available_products)} 个产品规格表")
            
            # 初始化产品分类字典，包含主要产品类型
            self.product_categories = {
                "NAS设备": {
                    "DS系列": [],
                    "RS系列": [],
                    "FS系列": [],
                    "SA系列": [],
                    "UC系列": [],
                },
                "存储扩充设备": {
                    "RX系列": [],
                    "DX系列": [],
                    "FX系列": [],
                    "RXD系列": [],
                },
                "PCIe扩充卡": {
                    "网卡": [],
                    "M.2转接卡": [],
                }
            }
            
            # 根据产品型号前缀分类主要产品
            print("\n正在根据产品型号前缀分类主要产品...")
            for product in self.available_products:
                # NAS设备分类
                if product.startswith('DS'):
                    self.product_categories["NAS设备"]["DS系列"].append(product)
                    print(f"  添加到DS系列: {product}")
                elif product.startswith('RS'):
                    self.product_categories["NAS设备"]["RS系列"].append(product)
                    print(f"  添加到RS系列: {product}")
                elif product.startswith('FS'):
                    self.product_categories["NAS设备"]["FS系列"].append(product)
                    print(f"  添加到FS系列: {product}")
                elif product.startswith('SA'):
                    self.product_categories["NAS设备"]["SA系列"].append(product)
                    print(f"  添加到SA系列: {product}")
                elif product.startswith('UC'):
                    self.product_categories["NAS设备"]["UC系列"].append(product)
                    print(f"  添加到UC系列: {product}")
                # 存储扩充设备分类
                elif product.startswith('RX') and not product.startswith('RXD'):
                    self.product_categories["存储扩充设备"]["RX系列"].append(product)
                    print(f"  添加到RX系列: {product}")
                elif product.startswith('DX'):
                    self.product_categories["存储扩充设备"]["DX系列"].append(product)
                    print(f"  添加到DX系列: {product}")
                elif product.startswith('FX'):
                    self.product_categories["存储扩充设备"]["FX系列"].append(product)
                    print(f"  添加到FX系列: {product}")
                elif product.startswith('RXD'):
                    self.product_categories["存储扩充设备"]["RXD系列"].append(product)
                    print(f"  添加到RXD系列: {product}")
                # PCIe扩充卡分类
                elif product.startswith('E10G'):
                    self.product_categories["PCIe扩充卡"]["网卡"].append(product)
                    print(f"  添加到网卡: {product}")
                elif product.startswith('M2D'):
                    self.product_categories["PCIe扩充卡"]["M.2转接卡"].append(product)
                    print(f"  添加到M.2转接卡: {product}")
            
            # 从硬盘分类查询表格读取存储设备分类
            print("\n正在从硬盘分类查询表格读取存储设备分类...")
            try:
                df_storage = pd.read_excel(SPECS_FILE, sheet_name="硬盘分类查询表格")
                if not df_storage.empty:
                    # 从列名中获取产品系列（排除'产品类型'和'产品系列'列）
                    storage_series = [col for col in df_storage.columns if col not in ['产品类型', '产品系列']]
                    if storage_series:
                        self.product_categories["存储设备"] = {series: [] for series in storage_series}
                        print(f"添加存储设备系列: {storage_series}")
                        
                        # 获取每列中非空的单元格内容作为产品型号
                        for series in storage_series:
                            if series in df_storage.columns:
                                # 获取该列所有非空值
                                models = df_storage[series].dropna().astype(str).tolist()
                                # 过滤掉空字符串和只包含空格的字符串
                                models = [model.strip() for model in models if model.strip()]
                                if models:
                                    self.product_categories["存储设备"][series] = models
                                    print(f"  添加存储设备型号到{series}: {models}")
            except Exception as e:
                print(f"读取硬盘分类查询表格时出错：{str(e)}")
            
            # 从配件分类查询表格读取配件类分类
            print("\n正在从配件分类查询表格读取配件类分类...")
            try:
                df_accessories = pd.read_excel(SPECS_FILE, sheet_name="配件分类查询表格")
                if not df_accessories.empty:
                    # 从列名中获取产品系列（排除'产品类型'和'产品系列'列）
                    accessory_series = [col for col in df_accessories.columns if col not in ['产品类型', '产品系列']]
                    if accessory_series:
                        self.product_categories["配件类"] = {series: [] for series in accessory_series}
                        print(f"添加配件类系列: {accessory_series}")
                        
                        # 获取每列中非空的单元格内容作为产品型号
                        for series in accessory_series:
                            if series in df_accessories.columns:
                                # 获取该列所有非空值
                                models = df_accessories[series].dropna().astype(str).tolist()
                                # 过滤掉空字符串和只包含空格的字符串
                                models = [model.strip() for model in models if model.strip()]
                                if models:
                                    self.product_categories["配件类"][series] = models
                                    print(f"  添加配件型号到{series}: {models}")
            except Exception as e:
                print(f"读取配件分类查询表格时出错：{str(e)}")
            
            # 对每个子类别中的产品型号进行排序
            for category in self.product_categories:
                for subcategory in self.product_categories[category]:
                    self.product_categories[category][subcategory].sort()
            
            # 更新全局变量
            global PRODUCT_CATEGORIES
            PRODUCT_CATEGORIES.clear()
            PRODUCT_CATEGORIES.update(self.product_categories)
            
            # 打印最终的产品分类统计
            self._print_category_statistics()
            
            # 确保下拉菜单数据正确更新
            if hasattr(self, 'category_combo'):
                categories = list(self.product_categories.keys())
                print(f"\n可选的产品类型: {categories}")
                self.category_combo['values'] = categories
            
            wb.close()
            
        except Exception as e:
            messagebox.showerror("错误", f"加载产品数据时出错：{str(e)}")
            self.available_products = []
            self.product_categories = {}
    
    def _process_category_data(self, df, fixed_category=None):
        """处理分类数据
        Args:
            df: DataFrame 包含分类数据的数据框
            fixed_category: str 固定的产品类型（可选）
        """
        if not df.empty:
            # 首先获取所有唯一的产品类型和系列
            unique_categories = {}
            for _, row in df.iterrows():
                category = fixed_category or str(row.get('产品类型', '')).strip()
                subcategory = str(row.get('产品系列', '')).strip()
                
                if category and not pd.isna(category):
                    if category not in unique_categories:
                        unique_categories[category] = set()
                    if subcategory and not pd.isna(subcategory):
                        unique_categories[category].add(subcategory)
            
            # 创建产品分类结构
            for category, subcategories in unique_categories.items():
                if category not in self.product_categories:
                    print(f"添加产品类型: {category}")
                    self.product_categories[category] = {}
                
                for subcategory in sorted(subcategories):
                    if subcategory not in self.product_categories[category]:
                        print(f"  添加产品系列: {category} - {subcategory}")
                        self.product_categories[category][subcategory] = []
            
            # 添加产品到对应的分类中
            for _, row in df.iterrows():
                category = fixed_category or str(row.get('产品类型', '')).strip()
                subcategory = str(row.get('产品系列', '')).strip()
                model = str(row.get('产品型号', '')).strip()
                
                if all(x and not pd.isna(x) for x in [category, subcategory, model]):
                    if model in self.available_products:  # 确保产品规格表存在
                        if model not in self.product_categories[category][subcategory]:
                            self.product_categories[category][subcategory].append(model)
                            print(f"  添加产品: {model} -> {category}/{subcategory}")
    
    def _auto_categorize_products(self):
        """根据产品型号前缀自动分类未分类的产品"""
        product_prefixes = {
            "DS": ("NAS设备", "DS系列"),
            "RS": ("NAS设备", "RS系列"),
            "FS": ("NAS设备", "FS系列"),
            "SA": ("NAS设备", "SA系列"),
            "UC": ("NAS设备", "UC系列"),
            "RX": ("存储扩充设备", "RX系列"),
            "DX": ("存储扩充设备", "DX系列"),
            "FX": ("存储扩充设备", "FX系列"),
            "RXD": ("存储扩充设备", "RXD系列"),
            "E10G": ("PCIe扩充卡", "网卡"),
            "M2D": ("PCIe扩充卡", "M.2转接卡")
        }
        
        # 检查每个产品是否已经被分类
        for product in self.available_products:
            already_categorized = False
            for category in self.product_categories.values():
                for products in category.values():
                    if product in products:
                        already_categorized = True
                        break
                if already_categorized:
                    break
            
            # 如果产品还未被分类，尝试使用前缀规则分类
            if not already_categorized:
                for prefix, (category, series) in product_prefixes.items():
                    if product.startswith(prefix):
                        if category in self.product_categories and series in self.product_categories[category]:
                            self.product_categories[category][series].append(product)
                            print(f"  自动分类产品: {product} -> {category}/{series}")
                        break
    
    def _print_category_statistics(self):
        """打印产品分类统计信息"""
        print("\n产品分类统计:")
        total_products = 0
        for category in self.product_categories:
            category_total = sum(len(products) for products in self.product_categories[category].values())
            total_products += category_total
            print(f"\n{category}: {category_total} 个产品")
            for subcategory, products in self.product_categories[category].items():
                print(f"  {subcategory}: {len(products)} 个产品")
                if products:  # 打印实际的产品型号
                    print(f"    产品型号: {', '.join(products)}")
        
        print(f"\n总计: {total_products} 个产品")
    
    def load_default_categories(self):
        """加载默认的产品分类"""
        # 根据产品型号前缀自动分类
        product_prefixes = {
            "DS": "DS系列",
            "RS": "RS系列",
            "FS": "FS系列",
            "SA": "SA系列",
            "UC": "UC系列",
            "RX": "RX系列",
            "DX": "DX系列",
            "FX": "FX系列",
            "RXD": "RXD系列",
            "E10G": "网卡",
            "M2D": "M.2转接卡"
        }
        
        # 重置产品分类
        PRODUCT_CATEGORIES.clear()
        PRODUCT_CATEGORIES.update({
            "NAS设备": {
                "DS系列": [],
                "RS系列": [],
                "FS系列": [],
                "SA系列": [],
                "UC系列": [],
            },
            "存储扩充设备": {
                "RX系列": [],
                "DX系列": [],
                "FX系列": [],
                "RXD系列": [],
            },
            "PCIe扩充卡": {
                "网卡": [],
                "M.2转接卡": [],
            },
            "硬盘": {
                "企业级硬盘": [],
                "数据中心硬盘": [],
                "监控级硬盘": [],
            },
            "内存": {
                "ECC内存": [],
                "非ECC内存": [],
            }
        })
        
        # 根据前缀将产品分类
        for product in self.available_products:
            for prefix, series in product_prefixes.items():
                if product.startswith(prefix):
                    # 确定产品类别
                    if prefix in ["DS", "RS", "FS", "SA", "UC"]:
                        category = "NAS设备"
                    elif prefix in ["RX", "DX", "FX", "RXD"]:
                        category = "存储扩充设备"
                    elif prefix in ["E10G", "M2D"]:
                        category = "PCIe扩充卡"
                    else:
                        continue
                    
                    # 将产品添加到对应的分类中
                    PRODUCT_CATEGORIES[category][series].append(product)
                    break
    
    def validate_product_data(self):
        """验证预定义的产品数据"""
        missing_products = []
        for category in PRODUCT_CATEGORIES:
            if category not in ["硬盘", "内存"]:  # 跳过硬盘和内存的验证
                for subcategory in PRODUCT_CATEGORIES[category]:
                    for product in PRODUCT_CATEGORIES[category][subcategory]:
                        if product not in self.available_products:
                            missing_products.append(product)
        
        if missing_products:
            messagebox.showwarning("警告", 
                f"以下产品在规格文件中未找到：\n{', '.join(missing_products)}\n" + 
                "这些产品将不会显示在选择列表中。")
    
    def setup_ui(self):
        """设置用户界面"""
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 客户信息区域
        customer_frame = ttk.LabelFrame(main_frame, text="客户信息", padding="5")
        customer_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        # 调整客户信息区域的布局和宽度
        ttk.Label(customer_frame, text="客户名称:").grid(row=0, column=0, padx=5)
        self.customer_name = ttk.Combobox(customer_frame, width=40)
        self.customer_name['values'] = list(self.existing_customers.keys())
        self.customer_name.grid(row=0, column=1, padx=5)
        # 设置为可编辑模式
        self.customer_name['state'] = 'normal'
        # 绑定事件
        self.customer_name.bind('<<ComboboxSelected>>', self.on_customer_selected)
        self.customer_name.bind('<KeyRelease>', self.on_customer_name_key_release)
        # 启用自动补全
        self.customer_name.bind('<Return>', self.on_customer_name_return)
        self.customer_name.bind('<Tab>', self.on_customer_name_return)
        # 设置下拉列表高度和显示
        self.customer_name.configure(height=10)
        # 绑定焦点事件 - 仅更新下拉列表值但不自动展开
        self.customer_name.bind('<FocusIn>', lambda e: self.update_customer_list())
        
        ttk.Label(customer_frame, text="联系人:").grid(row=0, column=2, padx=5)
        self.contact_name = ttk.Entry(customer_frame, width=30)
        self.contact_name.grid(row=0, column=3, padx=5)
        
        ttk.Label(customer_frame, text="联系电话:").grid(row=1, column=0, padx=5, pady=5)
        self.contact_phone = ttk.Entry(customer_frame, width=40)
        self.contact_phone.grid(row=1, column=1, padx=5)
        
        ttk.Label(customer_frame, text="报价日期:").grid(row=1, column=2, padx=5)
        self.quote_date = ttk.Entry(customer_frame, width=30)
        self.quote_date.grid(row=1, column=3, padx=5)
        self.quote_date.insert(0, datetime.now().strftime('%Y-%m-%d'))
        
        # 产品选择区域
        product_frame = ttk.LabelFrame(main_frame, text="产品选择", padding="5")
        product_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        # 第一行：产品类型和系列
        ttk.Label(product_frame, text="产品类型:").grid(row=0, column=0, padx=5)
        self.category_var = tk.StringVar()
        self.category_combo = ttk.Combobox(product_frame, textvariable=self.category_var, width=25, state="readonly")
        self.category_combo['values'] = list(self.product_categories.keys())  # 使用实例变量
        self.category_combo.grid(row=0, column=1, padx=5)
        self.category_combo.bind('<<ComboboxSelected>>', self.on_category_selected)
        
        ttk.Label(product_frame, text="产品系列:").grid(row=0, column=2, padx=5)
        self.subcategory_var = tk.StringVar()
        self.subcategory_combo = ttk.Combobox(product_frame, textvariable=self.subcategory_var, width=25, state="readonly")
        self.subcategory_combo.grid(row=0, column=3, padx=5)
        self.subcategory_combo.bind('<<ComboboxSelected>>', self.on_subcategory_selected)
        
        # 第二行：产品型号、数量和单价
        ttk.Label(product_frame, text="产品型号:").grid(row=1, column=0, padx=5, pady=5)
        self.product_var = tk.StringVar()
        self.product_combo = ttk.Combobox(product_frame, textvariable=self.product_var, width=25, state="readonly")
        self.product_combo.grid(row=1, column=1, padx=5, pady=5)
        self.product_combo.bind('<<ComboboxSelected>>', self.on_product_selected)
        
        ttk.Label(product_frame, text="数量:").grid(row=1, column=2, padx=5, pady=5)
        self.quantity_var = tk.StringVar(value="1")
        self.quantity_entry = ttk.Entry(product_frame, textvariable=self.quantity_var, width=10)
        self.quantity_entry.grid(row=1, column=3, padx=5, pady=5)
        
        ttk.Label(product_frame, text="单价:").grid(row=1, column=4, padx=5, pady=5)
        self.price_var = tk.StringVar()
        self.price_entry = ttk.Entry(product_frame, textvariable=self.price_var, width=15)
        self.price_entry.grid(row=1, column=5, padx=5, pady=5)
        
        ttk.Label(product_frame, text="折扣(%):").grid(row=1, column=6, padx=5, pady=5)
        self.discount_var = tk.StringVar(value="0")
        self.discount_entry = ttk.Entry(product_frame, textvariable=self.discount_var, width=8)
        self.discount_entry.grid(row=1, column=7, padx=5, pady=5)
        
        # 添加产品按钮
        add_btn = ttk.Button(product_frame, text="添加到报价单", command=self.add_product)
        add_btn.grid(row=1, column=8, padx=10, pady=5)
        
        # 报价项目列表
        list_frame = ttk.LabelFrame(main_frame, text="报价项目列表", padding="5")
        list_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # 调整表格列宽
        columns = ('序号', '产品型号', '规格描述', '数量', '单价', '折扣(%)', '折后价', '合计')
        self.tree = ttk.Treeview(list_frame, columns=columns, show='headings')
        
        # 设置列宽度 - 优化后所有列可完整显示
        column_widths = {
            '序号': 60,
            '产品型号': 150,
            '规格描述': 400,  # 减少宽度为其他列腾出空间
            '数量': 50,
            '单价': 80,
            '折扣(%)': 60,
            '折后价': 80,
            '合计': 80
        }
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=column_widths[col])  # 设置每列的宽度
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="删除选中项", command=self.delete_selected).grid(row=0, column=0, padx=10)
        ttk.Button(button_frame, text="生成报价单", command=self.generate_quote).grid(row=0, column=1, padx=10)
        ttk.Button(button_frame, text="清空列表", command=self.clear_list).grid(row=0, column=2, padx=10)
        ttk.Button(button_frame, text="清空所有输入", command=self.reset_inputs).grid(row=0, column=3, padx=10)
        
        # 配置grid权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        
    def center_window(self):
        """将窗口居中显示"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def on_customer_selected(self, event):
        """客户选择事件处理"""
        selected_customer = self.customer_name.get()
        if selected_customer in self.existing_customers:
            customer_info = self.existing_customers[selected_customer]
            self.contact_name.delete(0, tk.END)
            self.contact_name.insert(0, customer_info['contact'])
            self.contact_phone.delete(0, tk.END)
            self.contact_phone.insert(0, customer_info['phone'])
    
    def on_category_selected(self, event):
        """产品类型选择事件处理"""
        category = self.category_var.get()
        if category and category in self.product_categories:
            # 更新子类型下拉框
            subcategories = list(self.product_categories[category].keys())
            self.subcategory_combo['values'] = subcategories
            self.subcategory_combo.set('')  # 清空选择
            self.product_combo.set('')      # 清空产品选择
            self.product_combo['values'] = []  # 清空产品列表
        else:
            self.subcategory_combo['values'] = []
            self.product_combo['values'] = []
    
    def on_subcategory_selected(self, event):
        """产品子类型选择事件处理 - 按需加载产品数据"""
        category = self.category_var.get()
        subcategory = self.subcategory_var.get()
        if not (category and subcategory and 
               category in self.product_categories and 
               subcategory in self.product_categories[category]):
            self.product_combo['values'] = []
            self.product_combo.set('')
            return
            
        try:
            # 按需加载该子类别的产品数据
            if not os.path.exists(SPECS_FILE):
                messagebox.showerror("错误", f"未找到产品规格文件：{SPECS_FILE}")
                return
                
            # 根据产品系列前缀筛选产品
            prefix_map = {
                "DS系列": "DS",
                "RS系列": "RS", 
                "FS系列": "FS",
                "SA系列": "SA",
                "UC系列": "UC",
                "RX系列": "RX",
                "DX系列": "DX",
                "FX系列": "FX",
                "RXD系列": "RXD",
                "网卡": "E10G",
                "M.2转接卡": "M2D"
            }
            
            prefix = prefix_map.get(subcategory, "")
            if not prefix:
                self.product_combo['values'] = []
                return
                
            # 只读取工作表名而不加载整个文件
            wb = load_workbook(SPECS_FILE, read_only=True)
            products = [sheet for sheet in wb.sheetnames 
                       if sheet.startswith(prefix) and 
                       sheet not in ["产品汇总表", "查询表格", "硬盘分类查询表格", "配件分类查询表格"]]
            wb.close()
            
            # 更新产品下拉框
            if products:
                self.product_combo['values'] = sorted(products)
                self.product_combo.set('')
            else:
                self.product_combo['values'] = []
                self.product_combo.set('')
                messagebox.showwarning("提示", f"未找到{category}-{subcategory}的产品数据")
                
            # 缓存已加载的产品列表
            self.product_categories[category][subcategory] = products
            
        except Exception as e:
            messagebox.showerror("错误", f"加载产品数据时出错：{str(e)}")
            self.product_combo['values'] = []
    
    def on_product_selected(self, event):
        """产品选择事件处理 - 从规格配置动态提取参数"""
        selected_product = self.product_var.get()
        if not selected_product:
            return
            
        try:
            # 检查规格文件是否存在
            if not os.path.exists(SPECS_FILE):
                messagebox.showerror("错误", 
                    f"未找到产品规格文件：{SPECS_FILE}\n"
                    f"请确保文件位于：{os.path.abspath(SPECS_FILE)}")
                return
            
            # 读取产品规格
            wb = load_workbook(SPECS_FILE, read_only=True)
            
            # 1. 获取产品系列前缀(如"DS"、"RS"等)
            import re
            series_prefix = re.match(r'^([A-Z]+)', selected_product).group(1)
            
            # 2. 从"规格配置"工作表获取需要提取的参数
            if "规格配置" not in wb.sheetnames:
                messagebox.showerror("错误", 
                    "未找到'规格配置'工作表\n"
                    "请在Excel文件中创建'规格配置'工作表，格式为：\n"
                    "A列: 产品系列前缀(如DS/RS/FS等)\n"
                    "B列: 需要提取的参数(逗号分隔)")
                self.current_specs = ""
                return
                
            config_sheet = wb["规格配置"]
            params_to_extract = []
            
            # 查找匹配产品系列的配置行
            for row in config_sheet.iter_rows(min_row=1, values_only=True):
                if row and row[0] and str(row[0]).strip() == series_prefix:
                    if row[1]:  # 确保B列有参数配置
                        params_to_extract = [p.strip() for p in str(row[1]).split(',') if p.strip()]
                        break
                    
            if not params_to_extract:
                messagebox.showerror("错误", 
                    f"未找到产品系列'{series_prefix}'的配置\n"
                    f"请在'规格配置'工作表中添加一行：\n"
                    f"A列: {series_prefix}\n"
                    f"B列: 需要提取的参数(逗号分隔)")
                self.current_specs = ""
                return
                
            # 3. 检查产品工作表是否存在
            if selected_product not in wb.sheetnames:
                available_sheets = "\n".join(wb.sheetnames)
                messagebox.showerror("错误", 
                    f"未找到产品'{selected_product}'的规格表\n"
                    f"可用工作表有:\n{available_sheets}")
                return
            
            # 4. 从产品工作表中提取指定的参数
            product_sheet = wb[selected_product]
            specs = []
            
            # 扫描工作表查找参数
            for param in params_to_extract:
                found = False
                for row in product_sheet.iter_rows(min_row=1, max_row=100, values_only=True):
                    if row and len(row) > 1 and row[1] and str(row[1]).strip() == param:
                        if len(row) > 2 and row[2]:  # 参数值在第3列
                            specs.append(f"{param}: {str(row[2]).strip()}")
                            found = True
                            break
                if not found:
                    specs.append(f"{param}: 未找到")
                    
            # 格式化规格描述 - 使用换行符分隔参数
            self.current_specs = "\n".join(specs) if specs else "未找到规格信息"
            
        except Exception as e:
            messagebox.showerror("错误", 
                f"读取规格配置时出错：{str(e)}\n"
                f"文件路径: {os.path.abspath(SPECS_FILE)}")
            self.current_specs = ""
        finally:
            if 'wb' in locals():
                wb.close()
    
    def add_product(self):
        """添加产品到报价单"""
        product = self.product_var.get()
        if not product:
            messagebox.showerror("错误", "请选择产品型号")
            return
        
        try:
            quantity = int(self.quantity_var.get())
            price = float(self.price_var.get())
            discount = float(self.discount_var.get())
            discount = max(0, min(100, discount))  # 限制在0-100%之间
        except ValueError:
            messagebox.showerror("错误", "请输入有效的数量、单价或折扣")
            return
        
        # 计算折扣后价格和合计金额
        discounted_price = price * (1 - discount/100)
        total = quantity * discounted_price
        
        # 添加到表格
        item_id = len(self.quote_items) + 1
        self.tree.insert('', 'end', values=(
            item_id,
            product,
            getattr(self, 'current_specs', ''),
            quantity,
            f"￥{price:,.2f}",
            f"{discount:.1f}%",
            f"￥{discounted_price:,.2f}",
            f"￥{total:,.2f}"
        ))
        
        # 保存项目数据
        self.quote_items.append({
            'id': item_id,
            'product': product,
            'specs': getattr(self, 'current_specs', ''),
            'quantity': quantity,
            'price': price,
            'discount': discount,
            'discounted_price': discounted_price,
            'total': total
        })
        
        # 清空输入
        self.price_var.set('')
        self.quantity_var.set('1')
        self.discount_var.set('0')
    
    def delete_selected(self):
        """删除选中的项目"""
        selected_item = self.tree.selection()
        if not selected_item:
            return
        
        # 获取选中项的ID
        item_id = self.tree.item(selected_item)['values'][0]
        
        # 从列表中删除
        self.quote_items = [item for item in self.quote_items if item['id'] != item_id]
        
        # 从表格中删除
        self.tree.delete(selected_item)
        
        # 重新编号
        self.renumber_items()
    
    def clear_list(self):
        """清空报价单列表"""
        if messagebox.askyesno("确认", "确定要清空所有项目吗？"):
            self.tree.delete(*self.tree.get_children())
            self.quote_items = []

    def reset_inputs(self):
        """清空所有输入框"""
        if messagebox.askyesno("确认", "确定要清空所有输入内容吗？"):
            # 清空客户信息
            self.customer_name.set('')
            self.contact_name.delete(0, tk.END)
            self.contact_phone.delete(0, tk.END)
            self.quote_date.delete(0, tk.END)
            self.quote_date.insert(0, datetime.now().strftime('%Y-%m-%d'))
            
            # 清空产品选择
            self.category_var.set('')
            self.subcategory_var.set('')
            self.product_var.set('')
            self.quantity_var.set('1')
            self.price_var.set('')

    def renumber_items(self):
        """重新为项目编号"""
        items = self.tree.get_children()
        for i, item in enumerate(items, 1):
            values = list(self.tree.item(item)['values'])
            values[0] = i
            self.tree.item(item, values=values)
            
            # 更新项目列表中的ID
            for quote_item in self.quote_items:
                if quote_item['id'] == values[0]:
                    quote_item['id'] = i
    
    def generate_quote(self):
        """生成报价单"""
        if not self.quote_items:
            messagebox.showerror("错误", "报价单是空的")
            return
        
        if not self.customer_name.get():
            messagebox.showerror("错误", "请输入客户名称")
            return
        
        try:
            # 生成报价单
            wb = Workbook()
            ws = wb.active
            ws.title = "报价单"
            
            # 添加logo图片
            logo_path = os.path.join(os.path.dirname(__file__), "产品图片", "logo.png")
            print(f"尝试加载logo图片，路径: {logo_path}")
            if os.path.exists(logo_path):
                try:
                    from openpyxl.drawing.image import Image
                    img = Image(logo_path)
                    print(f"成功加载图片，原始尺寸: {img.width}x{img.height}")
                    # 保持原始比例，限制最大宽度为120
                    max_width = 120
                    if img.width > max_width:
                        ratio = max_width / img.width
                        img.width = max_width
                        img.height = int(img.height * ratio)
                    print(f"调整后尺寸: {img.width}x{img.height}")
                    # 将图片添加到H1单元格(右上角)
                    ws.add_image(img, 'H1')
                    # 设置H1单元格背景色与标题区域一致
                    ws['H1'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                    print("图片已添加到报价单")
                except Exception as e:
                    print(f"添加图片时出错: {str(e)}")
            else:
                print(f"图片文件不存在: {logo_path}")
            
            # 设置标题区域
            ws.merge_cells('A1:G1')  # 缩小合并区域为A1-G1，给logo留出H1位置
            title_cell = ws['A1']
            title_cell.value = "群晖产品报价单"
            title_cell.font = Font(name='微软雅黑', size=18, bold=True, color='2F5597')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            ws.row_dimensions[1].height = 60  # 增加行高以适应logo
            
            # 添加客户信息
            ws['A2'] = "客户名称："
            ws['B2'] = self.customer_name.get()
            ws['B2'].font = Font(bold=True)
            ws['D2'] = "联系人："
            ws['E2'] = self.contact_name.get()
            ws['E2'].font = Font(bold=True)
            
            ws['A3'] = "联系电话："
            ws['B3'] = self.contact_phone.get()
            ws['B3'].font = Font(bold=True)
            ws['D3'] = "报价日期："
            ws['E3'] = self.quote_date.get()
            ws['E3'].font = Font(bold=True)
            
            
            # 设置表头
            headers = ['序号', '产品型号', '规格描述', '数量', '单价', '折扣(%)', '折后价', '合计']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col)  # 下移两行给页眉留空间
                cell.value = header
                cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
                cell.border = Border(left=Side(style='medium'), 
                                   right=Side(style='medium'), 
                                   top=Side(style='medium'), 
                                   bottom=Side(style='medium'))
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 添加数据
            total_amount = 0
            for row, item in enumerate(self.quote_items, 7):  # 调整起始行
                ws.cell(row=row, column=1, value=item['id'])
                ws.cell(row=row, column=2, value=item['product'])
                
                # 规格描述单元格特殊处理 - 自动换行和自适应行高
                specs_cell = ws.cell(row=row, column=3, value=item['specs'])
                specs_cell.alignment = Alignment(wrap_text=True, vertical='center')
                
                ws.cell(row=row, column=4, value=item['quantity'])
                ws.cell(row=row, column=5, value=f"￥{item['price']:,.2f}")
                ws.cell(row=row, column=6, value=f"{item['discount']:.1f}%")
                ws.cell(row=row, column=7, value=f"￥{item['discounted_price']:,.2f}")
                ws.cell(row=row, column=8, value=f"￥{item['total']:,.2f}")
                
                # 设置单元格格式
                for col in range(1, 9):  # 包括所有列(A-H)
                    cell = ws.cell(row=row, column=col)
                    cell.border = NORMAL_BORDER
                    if col != 3:  # 规格描述列已单独设置
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 自动调整行高 - 根据换行符数量计算行数
                line_count = len(str(item['specs']).split('\n'))
                ws.row_dimensions[row].height = 20 * line_count  # 每行20像素
                
                total_amount += item['total']
            
            # 添加合计行
            total_row = len(self.quote_items) + 7
            ws.merge_cells(f'A{total_row}:G{total_row}')
            total_cell = ws[f'A{total_row}']
            total_cell.value = "合计金额"
            total_cell.font = Font(size=14, bold=True)
            total_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            total_amount_cell = ws[f'H{total_row}']
            total_amount_cell.value = f"￥{total_amount:,.2f}"
            total_amount_cell.font = Font(size=14, bold=True, color='C00000')
            total_amount_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            # 添加备注区域
            note_row = total_row + 1
            ws.merge_cells(f'A{note_row}:F{note_row}')
            note_cell = ws[f'A{note_row}']
            note_cell.value = "备注：1. 以上价格含税；2. 交货周期7个工作日；3. 付款方式：电汇"
            note_cell.font = Font(size=10, italic=True)
            
            # 设置列宽
            ws.column_dimensions['A'].width = 8    # 序号
            ws.column_dimensions['B'].width = 15   # 产品型号
            ws.column_dimensions['C'].width = 40   # 规格描述
            ws.column_dimensions['D'].width = 8    # 数量
            ws.column_dimensions['E'].width = 15   # 单价
            ws.column_dimensions['F'].width = 10   # 折扣(%)
            ws.column_dimensions['G'].width = 15   # 折后价
            ws.column_dimensions['H'].width = 15   # 合计
            
            # 生成报价单文件名
            customer_name = self.customer_name.get().strip()
            timestamp = datetime.now().strftime('%Y%m%d%H%M')
            # 跨平台文件名处理
            invalid_chars = '<>:"/\\|?*'
            quote_file = f"群晖产品报价_{customer_name}_{timestamp}.xlsx"
            for char in invalid_chars:
                quote_file = quote_file.replace(char, '_')
            
            # 使用报价单输出目录
            quote_file_path = os.path.join(QUOTE_DIR, quote_file)
            
            # 保存报价单
            wb.save(quote_file_path)
            
            # 更新客户信息表
            self.update_customer_info(customer_name, total_amount, quote_file)
            
            messagebox.showinfo("成功", f"报价单已生成：{quote_file_path}")
            
        except Exception as e:
            messagebox.showerror("错误", f"生成报价单时出错：{str(e)}")
    
    def update_customer_info(self, customer_name, total_amount, quote_file):
        """更新客户信息表"""
        try:
            # 如果文件存在，读取现有数据；否则创建新文件
            if os.path.exists(CUSTOMER_INFO_FILE):
                try:
                    wb = load_workbook(CUSTOMER_INFO_FILE)
                    ws = wb.active
                except Exception:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "客户信息"
                    # 设置表头
                    headers = ['客户名称', '联系人', '联系电话', '报价时间', '报价总额', '产品清单', '报价单文件']
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "客户信息"
                # 设置表头
                headers = ['客户名称', '联系人', '联系电话', '报价时间', '报价总额', '产品清单', '报价单文件']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
            
            # 获取下一个空行
            next_row = ws.max_row + 1
            
            # 准备产品清单
            products = []
            for item in self.quote_items:
                products.append(f"{item['product']}×{item['quantity']}")
            product_list = ", ".join(products)
            
            # 添加新记录
            ws.cell(row=next_row, column=1, value=customer_name)
            ws.cell(row=next_row, column=2, value=self.contact_name.get())
            ws.cell(row=next_row, column=3, value=self.contact_phone.get())
            ws.cell(row=next_row, column=4, value=self.quote_date.get())
            ws.cell(row=next_row, column=5, value=f"￥{total_amount:,.2f}")
            ws.cell(row=next_row, column=6, value=product_list)
            ws.cell(row=next_row, column=7, value=os.path.join(QUOTE_DIR, quote_file))
            
            # 调整列宽
            ws.column_dimensions['A'].width = 20  # 客户名称
            ws.column_dimensions['B'].width = 15  # 联系人
            ws.column_dimensions['C'].width = 15  # 联系电话
            ws.column_dimensions['D'].width = 12  # 报价时间
            ws.column_dimensions['E'].width = 15  # 报价总额
            ws.column_dimensions['F'].width = 50  # 产品清单
            ws.column_dimensions['G'].width = 40  # 报价单文件
            
            # 保存文件
            wb.save(CUSTOMER_INFO_FILE)
            
        except Exception as e:
            messagebox.showerror("错误", f"更新客户信息表时出错：{str(e)}")
    
    def on_customer_name_key_release(self, event):
        """智能客户名称输入辅助"""
        # 忽略特殊键和组合键
        if event.keysym in ['Up', 'Down', 'Left', 'Right', 'Return', 'Tab'] or \
           event.state & 0x0004:  # 忽略Ctrl/Alt等修饰键
            return
            
        current_text = self.customer_name.get()
        cursor_pos = self.customer_name.index(tk.INSERT)
        
        # 智能推荐逻辑
        if current_text:
            # 1. 精确匹配优先
            exact_matches = [c for c in self.existing_customers 
                           if c.lower() == current_text.lower()]
            
            # 2. 开头匹配次优
            start_matches = [c for c in self.existing_customers 
                           if c.lower().startswith(current_text.lower())]
            
            # 3. 包含匹配
            contains_matches = [c for c in self.existing_customers 
                              if current_text.lower() in c.lower()]
            
            # 合并结果，去重并保持优先级
            all_matches = list(dict.fromkeys(exact_matches + start_matches + contains_matches))
            
            # 限制推荐数量避免界面卡顿
            max_suggestions = 10
            if len(all_matches) > max_suggestions:
                all_matches = all_matches[:max_suggestions]
            
            self.customer_name['values'] = all_matches
            
            # 有匹配项时自动显示下拉框
            if all_matches:
                self.customer_name.event_generate('<Down>')
        else:
            # 清空时显示最近5个客户
            recent_customers = list(self.existing_customers.keys())[-5:]
            self.customer_name['values'] = recent_customers
        
        # 恢复光标位置并保持输入焦点
        self.after_idle(lambda: self.customer_name.icursor(cursor_pos))
        self.customer_name.focus_set()
    
    def on_customer_name_return(self, event):
        """处理客户名称输入框的回车和Tab事件"""
        current_text = self.customer_name.get().strip()
        if current_text:
            # 查找完全匹配的客户
            exact_match = next((customer for customer in self.existing_customers.keys() 
                              if customer.lower() == current_text.lower()), None)
            if exact_match:
                # 如果找到完全匹配，使用该客户信息
                self.customer_name.set(exact_match)
                self.on_customer_selected(None)
            else:
                # 查找部分匹配的客户
                matching_customers = [
                    customer for customer in self.existing_customers.keys()
                    if current_text.lower() in customer.lower()
                ]
                if len(matching_customers) == 1:
                    # 如果只有一个匹配项，使用该客户信息
                    self.customer_name.set(matching_customers[0])
                    self.on_customer_selected(None)
                elif len(matching_customers) > 1:
                    # 如果有多个匹配项，显示下拉列表
                    self.customer_name['values'] = matching_customers
                    self.customer_name.event_generate('<Down>')
        return 'break'  # 阻止默认的Tab行为
    
    def update_customer_list(self):
        """更新客户列表下拉框的内容"""
        current_text = self.customer_name.get().strip().lower()
        if current_text:
            matching_customers = [
                customer for customer in self.existing_customers.keys()
                if current_text in customer.lower()
            ]
            if matching_customers:
                self.customer_name['values'] = matching_customers
            else:
                self.customer_name['values'] = list(self.existing_customers.keys())
        else:
            self.customer_name['values'] = list(self.existing_customers.keys())
    
    def run(self):
        """运行程序"""
        self.root.mainloop()

if __name__ == "__main__":
    app = QuoteGenerator()
    app.run()
